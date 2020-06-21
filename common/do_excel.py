# _*_coding:utf-8_*_
# Author : wanghua
# Time   : 2019/12/13 22:30
# File   : do_excel.py
# IDE    : PyCharm


from openpyxl import load_workbook

class Cases:
    # 专门用来存放测试数据

    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None




class DoExcel:
    "此类用来读取excel里的测试数据，并返回测试结果到excel中"

    def __init__(self,excel_name):  #后面的函数都需要传这个变量，所以定义一个初始化函数
        self.excel_name=excel_name #指定操作的文件
        #self.sheet_name=sheet_name
        self.wb=load_workbook(excel_name) #实例化一个workbook实例,作为类属性





    def read_data(self,sheet_name):
        '''读取数据'''
        sheet=self.wb[sheet_name] #根据sheet_name定位到表单

        cases=[]

        for i in range(2,sheet.max_row+1):
            row_case = Cases() #实例化一个Cases对象
              #读取excel中的值，一行的数据赋值给一个row_case对象
            row_case.case_id=sheet.cell(row=i,column=1).value
            row_case.title=sheet.cell(row=i,column=2).value
            row_case.url = sheet.cell(row=i, column=3).value
            row_case.data = sheet.cell(row=i, column=4).value
            row_case.method = sheet.cell(row=i, column=5).value
            row_case.expected = sheet.cell(row=i, column=6).value
            if type(row_case.expected)==int:
                row_case.expected = str (row_case.expected) # excel取出的数值是int类型，与response返回结果里的内容（str）无法对比
            cases.append(row_case) #将每一行数据存入一个对象，将每一个对象存放在cases中

        return cases



    def write_data(self,sheet_name,row,actual,result):
        '''写入数据'''
        sheet=self.wb[sheet_name]
        sheet.cell(row,7).value=actual #写入实际结果
        sheet.cell(row,8).value=result #写入执行结果
        self.wb.save(self.excel_name) #写完数据保存



if __name__ == '__main__':
    from common import contants
    from common.request_method import RequestMethod
    do_excel=DoExcel(contants.case_file)  #利用.实现上下级，来引入其他目录下的文件
    cases=do_excel.read_data('login')
    #print(cases)
    req=RequestMethod()#实例化Reqeust类的对象
    for case in cases:
        resp= req.request_method(case.method,case.url,case.data) #返回请求结果
        print("case.expected:{}".format(case.expected))
        if resp.text==case.expected:
            do_excel.write_data('login',case.case_id+1,resp.text,'PASS')
        else:
            do_excel.write_data('login',case.case_id+1, resp.text, 'FAIL')







# _*_coding:utf-8_*_
# Author : wanghua
# Time   : 2020/3/25 09:33
# File   : doexcel.py
# IDE    : PyCharm

from openpyxl import load_workbook

class Cases:
    #专门存放数据的类

    def __init__(self):

        self.case_id= None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None


class DoExcel:
    # 读取数据的类，并将结果返回到excel中

    def __init__(self,excel_name):
        self.excel_name=excel_name#指定操作的文件
        self.wb=load_workbook(excel_name) #实例化workbook对象，定义为类属性

    def read_data(self,sheet_name):
        sheet=self.wb[sheet_name] #定位到具体操作的表单

        # 方法一：将所有测试数据存放在一个子列表中
        # testdata =[] #定义一个空的列表testdata来接收测试数据

        # for i in range(2,sheet.max_row+1):
        #     testdata.append(sheet.cell(row=i,column=1).value) # 取case_id值
        #     testdata.append(sheet.cell(row=i, column=2).value)  # 取title值
        #     testdata.append(sheet.cell(row=i, column=3).value)# 取url值
        #     testdata.append(sheet.cell(row=i, column=4).value)# 取data据
        #     testdata.append(sheet.cell(row=i, column=5).value)# 取method值
        #     testdata.append(sheet.cell(row=i, column=6).value)# 取expected值
        #
        # return testdata

        # 方法二：将所有测试数据存放在一个列表里，每一组的测试数据存放在列表的嵌套子字典里

        # testdata = []  # 定义一个空的列表testdata来接收测试数据
        #
        # for i in range(2,sheet.max_row+1):
        #     row_data={}
        #     row_data['case_id']=sheet.cell(row=i, column=1).value # 取case_id值
        #     row_data['title']=sheet.cell(row=i, column=2).value # 取title值
        #     row_data['url']=sheet.cell(row=i, column=3).value# 取url值
        #     row_data['data']=sheet.cell(row=i, column=4).value# 取data据
        #     row_data['method']=sheet.cell(row=i, column=5).value# 取method值
        #     row_data['expected']=sheet.cell(row=i, column=6).value# 取expected值
        #     testdata.append(row_data)
        # return testdata


        # 方法三：根据类与对象的思想,将每一条用例的数据存放到一个对象中，数据变为对象的属性，执行用例的时候，用对象去调用这些属性
        testdata=[]
        for i in range(2,sheet.max_row+1):
            row_case = Cases()  # 实例化Case类对象
            row_case.case_id = sheet.cell(row=i, column=1).value  # 取case_id值
            row_case.title=sheet.cell(row=i, column=2).value # 取title值
            row_case.url=sheet.cell(row=i, column=3).value# 取url值
            row_case.data=sheet.cell(row=i, column=4).value# 取data据
            row_case.method=sheet.cell(row=i, column=5).value# 取method值
            row_case.expected=sheet.cell(row=i, column=6).value# 取expected值

            if type(row_case.expected)== int :
                row_case.expected=str(row_case.expected)

            testdata.append(row_case)

        return testdata

    def write_result(self,sheet_name,row,actual,result):
        sheet=self.wb[sheet_name]
        sheet.cell(row=row,column=7).value=actual #写入实际结果
        sheet.cell(row=row,column=8).value=result  #写入执行结果
        self.wb.save(filename=self.excel_name)  #关闭表单文件生效


if __name__=='__main__':
    from common import contants
    from common import requestmethod
    print(contants.case_file)
    do_excel=DoExcel(contants.case_file)
    cases=do_excel.read_data('login')
    #print(cases)
    request=requestmethod.RequestMethod()  #实例化一个对象
    for case in cases:
        resp=request.request_method(case.method,case.url,case.data) #返回请求结果
        if resp.text == case.expected:
            do_excel.write_result("login",case.case_id+1,resp.text,"PASS")
        else:
            do_excel.write_result("login", case.case_id+1, resp.text, "Fail")




